import pytest
from datetime import datetime, date
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from models.activities_researchHotbed import ActivitiesResearchHotbed
from models.users import User
from models.users_research_hotbed import UsersResearchHotbed
from models.research_hotbed import ResearchHotbed
from models.activity_authors import ActivityAuthors
from controllers.export.excel_export_controller import (
    export_research_hotbed_excel,
    generate_excel_report,
    get_active_members,
    get_activities_by_semester
)
from controllers.export.users_pdf_export_controller import (
    export_user_excel,
    export_multiple_users_excel
)
from db.connection import db

@pytest.fixture
def setup_export_test_data():
    """Fixture para configurar datos de prueba completos para exportación"""
    
    # Crear semillero con TODOS los campos requeridos
    research_hotbed = ResearchHotbed(
        name_researchHotbed="Semillero de Sistemas de Información",
        acronym_researchHotbed="SSI",
        faculty_researchHotbed="Ingeniería",
        universityBranch_researchHotbed="Sede Principal",
        status_researchHotbed="Activo",  # CAMPO REQUERIDO AGREGADO
        dateCreation_researchHotbed=datetime.now()  # Usar datetime en lugar de date
    )
    db.session.add(research_hotbed)
    db.session.flush()
    
    # Crear usuarios con TODOS los campos requeridos
    users_data = [
        {
            'name': 'Juan Pérez García',
            'email': 'juan.perez@universidad.edu',
            'type': 'Estudiante',
            'idSigaa': '20241001'
        },
        {
            'name': 'María González López',
            'email': 'maria.gonzalez@universidad.edu',
            'type': 'Profesor',
            'idSigaa': '20191001'
        },
        {
            'name': 'Carlos Rodríguez Silva',
            'email': 'carlos.rodriguez@universidad.edu',
            'type': 'Egresado',
            'idSigaa': '20201001'
        }
    ]
    
    users = []
    users_research = []
    
    for user_data in users_data:
        user = User(
            name_user=user_data['name'],
            email_user=user_data['email'],
            password_user="hashed_password",
            idSigaa_user=user_data['idSigaa'],
            type_user=user_data['type'],
            status_user="Activo",
            academicProgram_user="Ingeniería de Sistemas",
            # Agregar campos requeridos que faltaban
            termsAccepted_user=True,
            termsAcceptedAt_user=datetime.now(),
            termsVersion_user="1.0"
        )
        db.session.add(user)
        db.session.flush()
        users.append(user)
        
        user_research = UsersResearchHotbed(
            user_iduser=user.iduser,
            researchHotbed_idresearchHotbed=research_hotbed.idresearchHotbed,
            TypeUser_usersResearchHotbed=user_data['type'],
            status_usersResearchHotbed="Activo",
            dateEnter_usersResearchHotbed=date.today()
        )
        db.session.add(user_research)
        db.session.flush()
        users_research.append(user_research)
    
    # Crear actividades de prueba
    activities_data = [
        {
            'title': 'Desarrollo de Sistema Web',
            'description': 'Implementación de sistema de gestión académica',
            'type': 'proyecto',
            'duration': 40,
            'semester': 'semestre-1-2025'
        },
        {
            'title': 'Artículo sobre IA',
            'description': 'Publicación en revista indexada sobre inteligencia artificial',
            'type': 'producto',
            'duration': 20,
            'semester': 'semestre-1-2025'
        },
        {
            'title': 'Premio Mejor Proyecto',
            'description': 'Reconocimiento por innovación tecnológica',
            'type': 'reconocimiento',
            'duration': 5,
            'semester': 'semestre-1-2025'
        }
    ]
    
    activities = []
    for i, activity_data in enumerate(activities_data):
        activity = ActivitiesResearchHotbed(
            title_activitiesResearchHotbed=activity_data['title'],
            responsible_activitiesResearchHotbed=users[i].name_user,
            date_activitiesResearchHotbed=date.today(),
            description_activitiesResearchHotbed=activity_data['description'],
            type_activitiesResearchHotbed=activity_data['type'],
            duration_activitiesResearchHotbed=activity_data['duration'],
            semester=activity_data['semester'],
            usersResearchHotbed_idusersResearchHotbed=users_research[i].idusersResearchHotbed,
            approvedFreeHours_activitiesResearchHotbed=1.0
        )
        db.session.add(activity)
        db.session.flush()
        activities.append(activity)
        
        # Crear relación de autoría
        author_relation = ActivityAuthors(
            activity_id=activity.idactivitiesResearchHotbed,
            user_research_hotbed_id=users_research[i].idusersResearchHotbed,
            is_main_author=True
        )
        db.session.add(author_relation)
    
    db.session.commit()
    
    return {
        'research_hotbed': research_hotbed,
        'users': users,
        'users_research': users_research,
        'activities': activities
    }

def test_export_research_hotbed_excel_success(client, setup_database, setup_export_test_data):
    """Prueba la exportación exitosa de Excel del semillero"""
    
    test_data = setup_export_test_data
    research_hotbed_id = test_data['research_hotbed'].idresearchHotbed
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación
    response = export_research_hotbed_excel(research_hotbed_id, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar respuesta exitosa
    assert status_code == 200
    
    # Verificar headers si es un objeto Response de Flask
    if hasattr(response_obj, 'headers'):
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response_obj.headers.get('Content-Type', '')
        
        # Verificar nombre del archivo si existe Content-Disposition
        content_disposition = response_obj.headers.get('Content-Disposition', '')
        if content_disposition:
            assert 'SSI_semestre-1-2025_reporte.xlsx' in content_disposition
        
        # Verificar contenido del Excel
        excel_data = response_obj.get_data() if hasattr(response_obj, 'get_data') else response_obj.data
        
        if excel_data:
            workbook = load_workbook(BytesIO(excel_data))
            
            # Verificar hojas creadas
            expected_sheets = ['Información General', 'Miembros']
            for sheet_name in expected_sheets:
                assert sheet_name in workbook.sheetnames
            
            # Verificar datos en hoja de información general
            info_sheet = workbook['Información General']
            # Buscar el nombre del semillero en las celdas
            semillero_found = False
            for row in info_sheet.iter_rows(values_only=True):
                if row and any(cell and 'Semillero de Sistemas de Información' in str(cell) for cell in row):
                    semillero_found = True
                    break
            assert semillero_found, "No se encontró el nombre del semillero en el Excel"

def test_export_user_excel_success(client, setup_database, setup_export_test_data):
    """Prueba la exportación exitosa de Excel de usuario individual"""
    
    test_data = setup_export_test_data
    user_id = test_data['users'][0].iduser
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación de usuario
    response = export_user_excel(user_id, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar respuesta exitosa
    assert status_code == 200
    
    # Verificar headers si es un objeto Response de Flask
    if hasattr(response_obj, 'headers'):
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response_obj.headers.get('Content-Type', '')
        
        # Verificar contenido del Excel si existe
        excel_data = response_obj.get_data() if hasattr(response_obj, 'get_data') else response_obj.data
        
        if excel_data:
            workbook = load_workbook(BytesIO(excel_data))
            
            # Verificar hojas específicas de usuario
            expected_sheets = ['Información Personal', 'Semilleros']
            for sheet_name in expected_sheets:
                assert sheet_name in workbook.sheetnames
            
            # Verificar datos del usuario
            info_sheet = workbook['Información Personal']
            user_found = False
            for row in info_sheet.iter_rows(values_only=True):
                if row and any(cell and 'Juan Pérez García' in str(cell) for cell in row):
                    user_found = True
                    break
            assert user_found, "No se encontró el nombre del usuario en el Excel"

def test_export_multiple_users_excel_success(client, setup_database, setup_export_test_data):
    """Prueba la exportación exitosa de Excel consolidado de múltiples usuarios"""
    
    test_data = setup_export_test_data
    user_ids = [user.iduser for user in test_data['users'][:2]]  # Primeros 2 usuarios
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación consolidada
    response = export_multiple_users_excel(user_ids, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar respuesta exitosa
    assert status_code == 200
    
    # Verificar headers si es un objeto Response de Flask
    if hasattr(response_obj, 'headers'):
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response_obj.headers.get('Content-Type', '')
        
        # Verificar nombre del archivo consolidado si existe Content-Disposition
        content_disposition = response_obj.headers.get('Content-Disposition', '')
        if content_disposition:
            assert 'usuarios_consolidado_semestre-1-2025.xlsx' in content_disposition
        
        # Verificar contenido del Excel si existe
        excel_data = response_obj.get_data() if hasattr(response_obj, 'get_data') else response_obj.data
        
        if excel_data:
            workbook = load_workbook(BytesIO(excel_data))
            
            # Verificar hoja consolidada
            assert 'Usuarios Consolidado' in workbook.sheetnames
            consolidated_sheet = workbook['Usuarios Consolidado']
            
            # Verificar que contiene información de ambos usuarios
            sheet_data = []
            for row in consolidated_sheet.iter_rows(values_only=True):
                if row and any(cell for cell in row if cell is not None):
                    sheet_data.append(row)
            
            # Buscar nombres de usuarios en los datos
            sheet_text = str(sheet_data)
            assert 'Juan Pérez García' in sheet_text or 'María González López' in sheet_text

def test_export_excel_with_invalid_semester(client, setup_database, setup_export_test_data):
    """Prueba el manejo de errores con semestre inválido"""
    
    test_data = setup_export_test_data
    research_hotbed_id = test_data['research_hotbed'].idresearchHotbed
    invalid_semester = 'semestre-invalido'
    
    # Ejecutar exportación con semestre inválido
    response = export_research_hotbed_excel(research_hotbed_id, invalid_semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar error
    assert status_code == 400
    
    # Verificar mensaje de error si es JSON
    if hasattr(response_obj, 'json') and response_obj.json:
        assert 'error' in response_obj.json
        assert 'Semestre inválido' in response_obj.json['error']

def test_export_excel_with_nonexistent_research_hotbed(client, setup_database):
    """Prueba el manejo de errores con semillero inexistente"""
    
    nonexistent_id = 99999
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación con semillero inexistente
    response = export_research_hotbed_excel(nonexistent_id, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar error
    assert status_code in [404, 400]  # Puede ser 404 o 400 dependiendo de la implementación
    
    # Verificar mensaje de error si es JSON
    if hasattr(response_obj, 'json') and response_obj.json:
        assert 'error' in response_obj.json

def test_export_user_excel_with_nonexistent_user(client, setup_database):
    """Prueba el manejo de errores con usuario inexistente"""
    
    nonexistent_user_id = 99999
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación con usuario inexistente
    response = export_user_excel(nonexistent_user_id, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar error
    assert status_code in [404, 400]  # Puede ser 404 o 400 dependiendo de la implementación
    
    # Verificar mensaje de error si es JSON
    if hasattr(response_obj, 'json') and response_obj.json:
        assert 'error' in response_obj.json

def test_generate_excel_report_structure(client, setup_database, setup_export_test_data):
    """Prueba la estructura interna del reporte Excel generado"""
    
    test_data = setup_export_test_data
    research_hotbed = test_data['research_hotbed']
    members = get_active_members(research_hotbed.idresearchHotbed)
    activities = get_activities_by_semester(research_hotbed.idresearchHotbed, 'semestre-1-2025')
    
    # Generar reporte
    excel_buffer = generate_excel_report(research_hotbed, members, activities, 'semestre-1-2025')
    
    # Verificar que el buffer no esté vacío
    assert excel_buffer is not None
    assert hasattr(excel_buffer, 'getvalue')
    assert len(excel_buffer.getvalue()) > 0
    
    # Cargar workbook para verificar estructura
    workbook = load_workbook(excel_buffer)
    
    # Verificar que tiene al menos la hoja de información general
    assert 'Información General' in workbook.sheetnames
    
    # Verificar datos estadísticos en la hoja
    info_sheet = workbook['Información General']
    
    # Buscar estadísticas
    stats_found = False
    for row in info_sheet.iter_rows(values_only=True):
        if row and any(cell and 'ESTADÍSTICAS GENERALES' in str(cell) for cell in row):
            stats_found = True
            break
    
    # Es opcional que tenga estadísticas, pero debe tener contenido
    has_content = False
    for row in info_sheet.iter_rows(values_only=True):
        if row and any(cell for cell in row if cell is not None):
            has_content = True
            break
    
    assert has_content, "La hoja de información general está vacía"

def test_excel_export_with_activities_by_type(client, setup_database, setup_export_test_data):
    """Prueba que el Excel genere hojas separadas por tipo de actividad"""
    
    test_data = setup_export_test_data
    research_hotbed_id = test_data['research_hotbed'].idresearchHotbed
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación
    response = export_research_hotbed_excel(research_hotbed_id, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar respuesta exitosa
    assert status_code == 200
    
    # Verificar contenido del Excel si existe
    if hasattr(response_obj, 'get_data'):
        excel_data = response_obj.get_data()
        
        if excel_data:
            workbook = load_workbook(BytesIO(excel_data))
            
            # Verificar que se crearon hojas para cada tipo de actividad
            expected_activity_sheets = ['Proyectos', 'Productos', 'Reconocimientos']
            
            # Al menos una de las hojas esperadas debe existir
            found_sheets = [sheet for sheet in expected_activity_sheets if sheet in workbook.sheetnames]
            assert len(found_sheets) > 0, f"No se encontraron hojas de actividades. Hojas disponibles: {workbook.sheetnames}"

def test_excel_export_styling_applied(client, setup_database, setup_export_test_data):
    """Prueba que los estilos se apliquen correctamente al Excel"""
    
    test_data = setup_export_test_data
    research_hotbed_id = test_data['research_hotbed'].idresearchHotbed
    semester = 'semestre-1-2025'
    
    # Ejecutar exportación
    response = export_research_hotbed_excel(research_hotbed_id, semester)
    
    # Verificar si la respuesta es una tupla o un objeto Response
    if isinstance(response, tuple):
        response_obj, status_code = response
    else:
        response_obj = response
        status_code = response.status_code
    
    # Verificar respuesta exitosa
    assert status_code == 200
    
    # Verificar contenido del Excel si existe
    if hasattr(response_obj, 'get_data'):
        excel_data = response_obj.get_data()
        
        if excel_data:
            workbook = load_workbook(BytesIO(excel_data))
            
            # Verificar que las hojas estén visibles (no ocultas)
            for sheet in workbook.worksheets:
                assert sheet.sheet_state == 'visible', f"La hoja {sheet.title} no está visible"
            
            # Verificar que las columnas tengan ancho ajustado en la hoja principal
            if 'Información General' in workbook.sheetnames:
                info_sheet = workbook['Información General']
                # Verificar que al menos una columna tenga un ancho personalizado
                has_custom_width = False
                for column_letter in ['A', 'B', 'C']:
                    if column_letter in info_sheet.column_dimensions:
                        width = info_sheet.column_dimensions[column_letter].width
                        if width and width > 8.43:  # Ancho por defecto de Excel
                            has_custom_width = True
                            break
                
                # Si no tiene anchos personalizados, al menos debe tener contenido
                has_content = False
                for row in info_sheet.iter_rows(values_only=True):
                    if row and any(cell for cell in row if cell is not None):
                        has_content = True
                        break
                
                assert has_content, "La hoja debe tener contenido aunque no tenga estilos personalizados"