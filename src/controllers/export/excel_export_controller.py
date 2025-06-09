from flask import jsonify, make_response
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from datetime import datetime
import logging

# Importar librerías para Excel
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from models.research_hotbed import ResearchHotbed
from models.users_research_hotbed import UsersResearchHotbed
from models.users import User
from models.activities_researchHotbed import ActivitiesResearchHotbed
from models.activity_authors import ActivityAuthors
from db.connection import db
from utils.semester_utils import format_semester_label_detailed, is_valid_semester

logger = logging.getLogger(__name__)

# ESTILOS GLOBALES ESTANDARIZADOS CON COLORES DE LA APP
def get_standard_styles():
    """Estilos estandarizados para todos los PDFs con colores de la aplicación"""
    styles = getSampleStyleSheet()
    
    return {
        'title': ParagraphStyle(
            'StandardTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1f2937'),
            fontName='Helvetica-Bold'
        ),
        'heading': ParagraphStyle(
            'StandardHeading',
            parent=styles['Heading1'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.HexColor('#7c2d92'),  # Morado de la app
            fontName='Helvetica-Bold'
        ),
        'subheading': ParagraphStyle(
            'StandardSubheading',
            parent=styles['Heading2'],
            fontSize=10,
            spaceAfter=8,
            textColor=colors.HexColor('#be185d'),  # Rosa de la app
            fontName='Helvetica-Bold'
        ),
        'normal': ParagraphStyle(
            'StandardNormal',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor('#374151')
        ),
        'small': ParagraphStyle(
            'StandardSmall',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#6b7280')
        )
    }

def get_standard_table_style():
    """Estilo estandarizado para tablas con colores de la app"""
    return TableStyle([
        # Header - Rosa/Morado degradado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d946ef')),  # Rosa vibrante
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fdf2f8')),  # Rosa muy claro
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#f3e8ff')),  # Lila claro
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8)
    ])

def get_info_table_style():
    """Estilo para tablas de información con colores de la app"""
    return TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#a21caf')),  # Morado medio
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#f3e8ff')),  # Lila claro
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fdf2f8')),  # Rosa muy claro
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
    ])

def get_stats_table_style():
    """Estilo para tablas de estadísticas con colores de la app"""
    return TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c026d3')),  # Rosa fuerte
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fae8ff')),  # Lila muy claro
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#be185d')),  # Rosa fuerte
        ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#be185d')),  # Rosa fuerte
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8)
    ])

def export_research_hotbed_excel(research_hotbed_id, semester):
    """
    Genera un archivo Excel completo del semillero para el semestre especificado
    con hojas separadas para cada tipo de información
    """
    try:
        # Validaciones
        if not semester or not is_valid_semester(semester):
            return jsonify({"error": "Semestre inválido"}), 400
            
        # Obtener datos del semillero
        research_hotbed = ResearchHotbed.query.get(research_hotbed_id)
        if not research_hotbed:
            return jsonify({"error": "Semillero no encontrado"}), 404
            
        # Obtener miembros activos
        members = get_active_members(research_hotbed_id)
        
        # Obtener actividades del semestre
        activities = get_activities_by_semester(research_hotbed_id, semester)
        
        # Generar Excel
        excel_buffer = generate_excel_report(research_hotbed, members, activities, semester)
        
        # Crear respuesta
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{research_hotbed.acronym_researchHotbed}_{semester}_reporte.xlsx"'
        
        logger.info(f"Excel generado exitosamente para semillero {research_hotbed_id}, semestre {semester}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando Excel: {str(e)}")
        return jsonify({"error": f"Error generando Excel: {str(e)}"}), 500

def create_general_info_sheet(writer, research_hotbed, members, activities, semester):
    """Crea la hoja de información general del semillero con formato mejorado"""
    
    # Información básica del semillero con mejor organización
    info_data = [
        ['INFORMACIÓN DEL SEMILLERO', ''],
        ['Nombre del Semillero', research_hotbed.name_researchHotbed],
        ['Acrónimo/Siglas', research_hotbed.acronym_researchHotbed],
        ['Facultad', research_hotbed.faculty_researchHotbed],
        ['Sede Universitaria', research_hotbed.universityBranch_researchHotbed],
        ['Período del Reporte', format_semester_label_detailed(semester)],
        ['Fecha de Generación', datetime.now().strftime('%d de %B de %Y - %H:%M hrs')],
        ['', ''],
        ['ESTADÍSTICAS GENERALES', ''],
        ['Total de Miembros Registrados', f'{len(members)} personas'],
        ['Miembros con Estado Activo', f'{len([m for m in members if m["status"] == "Activo"])} personas'],
        ['Miembros con Estado Inactivo', f'{len([m for m in members if m["status"] != "Activo"])} personas'],
        ['Total de Actividades Registradas', f'{len(activities)} actividades'],
        ['Horas Académicas Totales', f'{sum(a["duration"] for a in activities)} horas'],
        ['Actividades con Horas Aprobadas', f'{len([a for a in activities if a["approved_free_hours"]])} actividades'],
        ['Actividades Pendientes de Aprobación', f'{len([a for a in activities if not a["approved_free_hours"]])} actividades'],
        ['Proyectos de Investigación', f'{len([a for a in activities if "proyecto" in a["type"].lower()])} proyectos'],
        ['Productos Académicos', f'{len([a for a in activities if "producto" in a["type"].lower()])} productos'],
        ['Reconocimientos Obtenidos', f'{len([a for a in activities if "reconocimiento" in a["type"].lower()])} reconocimientos']
    ]
    
    df_info = pd.DataFrame(info_data, columns=['Concepto', 'Detalle'])
    df_info.to_excel(writer, sheet_name='Información General', index=False)

def create_members_sheet(writer, members):
    """Crea la hoja detallada de miembros con formato mejorado"""
    if not members:
        df_empty = pd.DataFrame({'Información': ['No se encontraron miembros registrados en este semillero.']})
        df_empty.to_excel(writer, sheet_name='Miembros', index=False)
        return
    
    # Separar y ordenar miembros
    active_members = [m for m in members if m['status'] == 'Activo']
    inactive_members = [m for m in members if m['status'] != 'Activo']
    sorted_members = active_members + inactive_members
    
    members_data = []
    for member in sorted_members:
        members_data.append({
            'Nombre Completo': member['name'],
            'Correo Electrónico': member['email'],
            'ID SIGAA': member['idSigaa'] if member['idSigaa'] else 'No registrado',
            'Tipo de Usuario': member['type'],
            'Estado Actual': member['status'],
            'Fecha de Ingreso': member['dateEnter'].strftime('%d/%m/%Y') if member['dateEnter'] else 'No registrada',
            'Fecha de Salida': member['dateExit'].strftime('%d/%m/%Y') if member['dateExit'] else 'N/A',
            'Observaciones': member['observation'] if member['observation'] else ('Sin observaciones' if member['status'] == 'Activo' else 'Sin detalles registrados')
        })
    
    df_members = pd.DataFrame(members_data)
    df_members.to_excel(writer, sheet_name='Miembros', index=False)

def create_activities_sheets(writer, activities):
    """Crea hojas separadas para cada tipo de actividad"""
    if not activities:
        df_empty = pd.DataFrame({'Mensaje': ['No hay actividades registradas para este semestre.']})
        df_empty.to_excel(writer, sheet_name='Actividades', index=False)
        return
    
    # Agrupar actividades por tipo
    activities_by_type = {}
    for activity in activities:
        activity_type = activity['type']
        if activity_type not in activities_by_type:
            activities_by_type[activity_type] = []
        activities_by_type[activity_type].append(activity)
    
    # Crear hoja para cada tipo
    for activity_type, type_activities in activities_by_type.items():
        sheet_name = get_sheet_name_for_type(activity_type)
        
        if 'proyecto' in activity_type.lower():
            create_projects_sheet(writer, type_activities, sheet_name)
        elif 'producto' in activity_type.lower():
            create_products_sheet(writer, type_activities, sheet_name)
        elif 'reconocimiento' in activity_type.lower():
            create_recognitions_sheet(writer, type_activities, sheet_name)
        else:
            create_generic_activities_sheet(writer, type_activities, sheet_name)

def get_sheet_name_for_type(activity_type):
    """Obtiene el nombre de la hoja según el tipo de actividad"""
    if 'proyecto' in activity_type.lower():
        return 'Proyectos'
    elif 'producto' in activity_type.lower():
        return 'Productos'
    elif 'reconocimiento' in activity_type.lower():
        return 'Reconocimientos'
    else:
        # Limpiar el nombre y limitar a 31 caracteres (límite de Excel)
        clean_name = activity_type.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        return clean_name.capitalize()[:31]

def generate_excel_report(research_hotbed, members, activities, semester):
    """Genera el reporte Excel con múltiples hojas"""
    buffer = BytesIO()
    
    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # 1. Hoja de información del semillero y miembros
            create_general_info_sheet(writer, research_hotbed, members, activities, semester)
            
            # 2. Hoja de miembros detallada
            create_members_sheet(writer, members)
            
            # 3. Hojas de actividades por tipo
            create_activities_sheets(writer, activities)
        
        # Recargar el workbook para aplicar estilos
        buffer.seek(0)
        workbook = load_workbook(buffer)
        
        # Verificar que tengamos al menos una hoja
        if len(workbook.sheetnames) == 0:
            # Crear una hoja vacía si no hay ninguna
            ws = workbook.create_sheet("Información General")
            ws['A1'] = "No hay datos disponibles"
        
        # Aplicar estilos
        apply_excel_styles(workbook)
        
        # Asegurar que al menos una hoja esté visible
        for sheet in workbook.worksheets:
            if sheet.sheet_state == 'hidden':
                sheet.sheet_state = 'visible'
        
        # Si todas están ocultas, hacer visible la primera
        if all(sheet.sheet_state == 'hidden' for sheet in workbook.worksheets):
            workbook.worksheets[0].sheet_state = 'visible'
        
        # Guardar cambios
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        logger.error(f"Error en generate_excel_report: {str(e)}")
        # Crear un Excel básico en caso de error
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            error_data = pd.DataFrame({
                'Error': [f'No se pudieron generar los datos: {str(e)}'],
                'Semillero': [research_hotbed.name_researchHotbed if research_hotbed else 'Desconocido'],
                'Semestre': [semester]
            })
            error_data.to_excel(writer, sheet_name='Error', index=False)
        
        buffer.seek(0)
        return buffer

# Modificar la función apply_excel_styles para ser más robusta
def apply_excel_styles(workbook):
    """Aplica estilos profesionales y formales al workbook"""
    
    try:
        # Colores formales y corporativos
        primary_blue = '1F4E79'      # Azul marino corporativo
        secondary_blue = '2E75B6'    # Azul medio
        light_blue = 'E7F3FF'       # Azul muy claro
        dark_gray = '2F2F2F'        # Gris oscuro
        medium_gray = '5A5A5A'      # Gris medio
        light_gray = 'F8F9FA'       # Gris muy claro
        white = 'FFFFFF'             # Blanco
        accent_green = '4A7C59'      # Verde corporativo
        
        # Estilos profesionales
        header_font = Font(bold=True, color=white, size=11, name='Calibri')
        subheader_font = Font(bold=True, color=dark_gray, size=10, name='Calibri')
        body_font = Font(color=dark_gray, size=10, name='Calibri')
        
        # Fills (fondos)
        header_fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type='solid')
        subheader_fill = PatternFill(start_color=secondary_blue, end_color=secondary_blue, fill_type='solid')
        body_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type='solid')
        alternate_fill = PatternFill(start_color=white, end_color=white, fill_type='solid')
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Bordes profesionales
        thin_border = Side(border_style='thin', color=medium_gray)
        medium_border = Side(border_style='medium', color=primary_blue)
        
        header_border = Border(
            left=thin_border, right=thin_border, 
            top=medium_border, bottom=medium_border
        )
        
        body_border = Border(
            left=thin_border, right=thin_border, 
            top=thin_border, bottom=thin_border
        )
        
        # Aplicar estilos a todas las hojas
        for sheet_name in workbook.sheetnames:
            try:
                sheet = workbook[sheet_name]
                
                # Asegurar que la hoja esté visible
                sheet.sheet_state = 'visible'
                
                # Configurar ancho de columnas de manera más inteligente
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                # Calcular longitud considerando saltos de línea
                                cell_length = max(len(str(line)) for line in str(cell.value).split('\n'))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # Ajustar ancho con límites más profesionales
                    if max_length == 0:
                        adjusted_width = 12
                    elif max_length < 8:
                        adjusted_width = 12
                    elif max_length < 20:
                        adjusted_width = max_length + 3
                    elif max_length < 40:
                        adjusted_width = max_length + 2
                    else:
                        adjusted_width = 45
                        
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar estilos específicos según el tipo de hoja
                if sheet_name == 'Información General':
                    apply_info_sheet_styles(sheet, header_font, subheader_font, body_font, 
                                          header_fill, body_fill, alternate_fill, 
                                          center_alignment, left_alignment, header_border, body_border)
                
                elif sheet_name == 'Miembros':
                    apply_members_sheet_styles(sheet, header_font, body_font, 
                                             header_fill, body_fill, alternate_fill,
                                             center_alignment, left_alignment, header_border, body_border)
                
                else:  # Hojas de actividades
                    apply_activities_sheet_styles(sheet, header_font, body_font,
                                                header_fill, body_fill, alternate_fill,
                                                center_alignment, left_alignment, header_border, body_border)
            except Exception as e:
                logger.error(f"Error aplicando estilos a la hoja {sheet_name}: {str(e)}")
                continue
                
    except Exception as e:
        logger.error(f"Error general aplicando estilos: {str(e)}")

def export_research_hotbed_pdf(research_hotbed_id, semester):
    """
    Genera un PDF completo del semillero para el semestre especificado
    """
    try:
        # Validaciones
        if not semester or not is_valid_semester(semester):
            return jsonify({"error": "Semestre inválido"}), 400
            
        # Obtener datos del semillero
        research_hotbed = ResearchHotbed.query.get(research_hotbed_id)
        if not research_hotbed:
            return jsonify({"error": "Semillero no encontrado"}), 404
            
        # Obtener miembros activos
        members = get_active_members(research_hotbed_id)
        
        # Obtener actividades del semestre
        activities = get_activities_by_semester(research_hotbed_id, semester)
        
        # Generar PDF
        pdf_buffer = generate_pdf_report(research_hotbed, members, activities, semester)
        
        # Crear respuesta
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{research_hotbed.acronym_researchHotbed}_{semester}_reporte.pdf"'
        
        logger.info(f"PDF generado exitosamente para semillero {research_hotbed_id}, semestre {semester}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando PDF: {str(e)}")
        return jsonify({"error": f"Error generando PDF: {str(e)}"}), 500

def get_active_members(research_hotbed_id):
    """Obtiene TODOS los miembros del semillero (activos e inactivos)"""
    try:
        # CAMBIO: Eliminar filtro de status para mostrar todos los miembros
        members_query = db.session.query(User, UsersResearchHotbed).join(
            UsersResearchHotbed, User.iduser == UsersResearchHotbed.user_iduser
        ).filter(
            UsersResearchHotbed.researchHotbed_idresearchHotbed == research_hotbed_id
            # Removido: UsersResearchHotbed.status_usersResearchHotbed == 'Activo'
        ).order_by(
            # Activos primero (Activo = 1, otros = 0), luego alfabético
            (UsersResearchHotbed.status_usersResearchHotbed == 'Activo').desc(),
            User.name_user.asc()
        ).all()
        
        return [{
            'name': user.name_user,
            'email': user.email_user,
            'idSigaa': user.idSigaa_user,
            'type': user_research.TypeUser_usersResearchHotbed,
            'dateEnter': user_research.dateEnter_usersResearchHotbed,
            'dateExit': user_research.dateExit_usersResearchHotbed, # Campo que SÍ existe
            'observation': user_research.observation_usersResearchHotbed, # Usar observation en lugar de exitReason
            'status': user_research.status_usersResearchHotbed
        } for user, user_research in members_query]
        
    except Exception as e:
        logger.error(f"Error obteniendo miembros: {str(e)}")
        return []

def get_activities_by_semester(research_hotbed_id, semester):
    """Obtiene actividades filtradas por semestre usando el campo 'semester' de la actividad"""
    try:
        # Consulta base de actividades filtradas por el campo semester
        activities_query = db.session.query(ActivitiesResearchHotbed).filter(
            ActivitiesResearchHotbed.usersResearchHotbed_idusersResearchHotbed.in_(
                db.session.query(UsersResearchHotbed.idusersResearchHotbed).filter(
                    UsersResearchHotbed.researchHotbed_idresearchHotbed == research_hotbed_id
                )
            ),
            ActivitiesResearchHotbed.semester == semester  # Filtrar por el campo semester
        ).all()
        
        # Procesar actividades
        filtered_activities = []
        for activity in activities_query:
            # Obtener autores
            authors_data = get_activity_authors(activity.idactivitiesResearchHotbed)
            
            activity_data = {
                'id': activity.idactivitiesResearchHotbed,
                'title': activity.title_activitiesResearchHotbed,
                'responsible': activity.responsible_activitiesResearchHotbed,
                'date': activity.date_activitiesResearchHotbed,
                'description': activity.description_activitiesResearchHotbed,
                'type': activity.type_activitiesResearchHotbed,
                'duration': activity.duration_activitiesResearchHotbed or 0,
                'start_time': activity.startTime_activitiesResearchHotbed,
                'end_time': activity.endTime_activitiesResearchHotbed,
                'approved_free_hours': bool(activity.approvedFreeHours_activitiesResearchHotbed),
                'reference_number': getattr(activity, 'reference_number', None),
                'publication_date': getattr(activity, 'publication_date', None),
                'organization_name': getattr(activity, 'organization_name', None),
                'authors': authors_data
            }
            
            # Obtener datos específicos según el tipo
            if (activity.type_activitiesResearchHotbed and 
                activity.type_activitiesResearchHotbed.lower() == 'proyecto' and 
                activity.projectsResearchHotbed_idprojectsResearchHotbed):
                
                from models.projects_researchHotbed import ProjectsResearchHotbed
                project = db.session.query(ProjectsResearchHotbed).filter_by(
                    idprojectsResearchHotbed=activity.projectsResearchHotbed_idprojectsResearchHotbed
                ).first()
                if project:
                    activity_data['project_data'] = {
                        'name': project.name_projectsResearchHotbed or 'Sin especificar',
                        'reference_number': project.referenceNumber_projectsResearchHotbed,
                        'start_date': project.startDate_projectsResearchHotbed,
                        'end_date': project.endDate_projectsResearchHotbed
                    }
            
            elif (activity.type_activitiesResearchHotbed and 
                  activity.type_activitiesResearchHotbed.lower() == 'producto' and 
                  activity.productsResearchHotbed_idproductsResearchHotbed):
                
                from models.products_researchHotbed import ProductsResearchHotbed
                product = db.session.query(ProductsResearchHotbed).filter_by(
                    idproductsResearchHotbed=activity.productsResearchHotbed_idproductsResearchHotbed
                ).first()
                if product:
                    activity_data['product_data'] = {
                        'category': product.category_productsResearchHotbed,
                        'type': product.type_productsResearchHotbed,
                        'description': product.description_productsResearchHotbed
                    }
            
            elif (activity.type_activitiesResearchHotbed and 
                  activity.type_activitiesResearchHotbed.lower() == 'reconocimiento' and 
                  activity.recognitionsResearchHotbed_idrecognitionsResearchHotbed):
                
                from models.recognitions_researchHotbed import RecognitionsResearchHotbed
                recognition = db.session.query(RecognitionsResearchHotbed).filter_by(
                    idrecognitionsResearchHotbed=activity.recognitionsResearchHotbed_idrecognitionsResearchHotbed
                ).first()
                if recognition:
                    activity_data['recognition_data'] = {
                        'project_name': recognition.projectName_recognitionsResearchHotbed,
                        'organization_name': recognition.organizationName_recognitionsResearchHotbed
                    }
            
            filtered_activities.append(activity_data)
                
        return filtered_activities
        
    except Exception as e:
        logger.error(f"Error obteniendo actividades: {str(e)}")
        return []

def get_activity_authors(activity_id):
    """Obtiene los autores de una actividad"""
    try:
        authors_query = db.session.query(User, ActivityAuthors).join(
            UsersResearchHotbed, ActivityAuthors.user_research_hotbed_id == UsersResearchHotbed.idusersResearchHotbed
        ).join(
            User, UsersResearchHotbed.user_iduser == User.iduser
        ).filter(
            ActivityAuthors.activity_id == activity_id
        ).all()
        
        main_authors = []
        co_authors = []
        
        for user, author_relation in authors_query:
            if author_relation.is_main_author:
                main_authors.append(user.name_user)
            else:
                co_authors.append(user.name_user)
                
        return {
            'main_authors': main_authors,
            'co_authors': co_authors
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo autores: {str(e)}")
        return {'main_authors': [], 'co_authors': []}

def generate_pdf_report(research_hotbed, members, activities, semester):
    """Genera el PDF usando estilos con colores de la app"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        topMargin=0.8*inch, 
        bottomMargin=0.8*inch, 
        leftMargin=0.7*inch, 
        rightMargin=0.7*inch
    )
    
    # Obtener estilos estandarizados
    std_styles = get_standard_styles()
    story = []
    
    # Título principal con colores de la app
    title = Paragraph("REPORTE DE SEMILLERO DE INVESTIGACIÓN", std_styles['title'])
    story.append(title)
    story.append(Spacer(1, 10))
    
    # Subtítulo con información del semillero en colores de la app
    semillero_info = Paragraph(
        f"<b><font color='#be185d'>{research_hotbed.name_researchHotbed} ({research_hotbed.acronym_researchHotbed})</font></b><br/>"
        f"<font color='#7c2d92'>{format_semester_label_detailed(semester)} | {research_hotbed.faculty_researchHotbed} | {research_hotbed.universityBranch_researchHotbed}</font>",
        std_styles['normal']
    )
    semillero_info.alignment = TA_CENTER
    story.append(semillero_info)
    story.append(Spacer(1, 20))
    
    # Información del semillero
    info_data = [
        ['Semillero:', research_hotbed.name_researchHotbed],
        ['Acrónimo:', research_hotbed.acronym_researchHotbed],
        ['Facultad:', research_hotbed.faculty_researchHotbed],
        ['Sede:', research_hotbed.universityBranch_researchHotbed],
        ['Semestre:', format_semester_label_detailed(semester)],
        ['Fecha de reporte:', datetime.now().strftime('%d/%m/%Y %H:%M')]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_table.setStyle(get_info_table_style())
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Resumen estadístico - ACTUALIZADO para mostrar activos e inactivos
    story.append(Paragraph("RESUMEN ESTADÍSTICO", std_styles['heading']))
    
    # Usar el mismo criterio de filtrado que en la tabla
    active_members = [m for m in members if m['status'] == 'Activo']
    inactive_members = [m for m in members if m['status'] != 'Activo']
    total_hours = sum(a['duration'] for a in activities)
    approved_activities = len([a for a in activities if a['approved_free_hours']])
    
    stats_data = [
        ['Total miembros', str(len(members)), 'Miembros activos', str(len(active_members))],
        ['Miembros inactivos', str(len(inactive_members)), 'Horas totales', f'{total_hours}h'],
        ['Total actividades', str(len(activities)), 'Actividades aprobadas', str(approved_activities)],
        ['Proyectos', str(len([a for a in activities if 'proyecto' in a['type'].lower()])), 'Productos', str(len([a for a in activities if 'producto' in a['type'].lower()]))],
        ['Reconocimientos', str(len([a for a in activities if 'reconocimiento' in a['type'].lower()])), 'Pendientes', str(len(activities) - approved_activities)]
    ]
    
    stats_table = Table(stats_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch])
    stats_table.setStyle(get_stats_table_style())
    story.append(stats_table)
    story.append(Spacer(1, 20))
    
    # Tabla de miembros - MOSTRAR TODOS CON STATUS, FECHA DE SALIDA Y OBSERVACIÓN
    story.append(Paragraph("MIEMBROS DEL SEMILLERO", std_styles['heading']))
    
    if members:
        # Separar activos e inactivos para mostrar activos primero
        active_members = [m for m in members if m['status'] == 'Activo']
        inactive_members = [m for m in members if m['status'] != 'Activo']
        sorted_members = active_members + inactive_members
        
        # Incluir columnas para fecha de salida y observación
        members_data = [['Nombre', 'ID SIGAA', 'Tipo', 'Estado', 'F. Ingreso', 'F. Salida', 'Observación']]
        
        for member in sorted_members:
            name = member['name'][:18] + '...' if len(member['name']) > 18 else member['name']
            type_user = member['type'][:8] + '...' if len(member['type']) > 8 else member['type']
            
            # Formatear fechas
            date_enter = member['dateEnter'].strftime('%m/%Y') if member['dateEnter'] else 'N/A'
            date_exit = member['dateExit'].strftime('%m/%Y') if member['dateExit'] else '-'
            
            # Formatear observación
            observation = ''
            if member['observation']:
                observation = member['observation'][:12] + '...' if len(member['observation']) > 12 else member['observation']
            else:
                observation = '-' if member['status'] == 'Activo' else 'Sin observación'
            
            members_data.append([
                name,
                member['idSigaa'] or 'N/A',
                type_user,
                member['status'],
                date_enter,
                date_exit,
                observation
            ])
        
        # Ajustar anchos de columna para incluir nuevos campos
        # Reducir algunos anchos para hacer espacio para fecha de salida y observación
        members_table = Table(members_data, colWidths=[1.4*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.6*inch, 1.0*inch])
        
        # Estilo con colores diferentes para activos/inactivos
        table_style = get_standard_table_style()
        
        # Agregar colores especiales para miembros inactivos
        for row_idx, member in enumerate(sorted_members, 1):  # Saltar header, empezar desde 1
            if member['status'] != 'Activo':  # Si no es activo
                table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#fee2e2'))  # Fondo rojo claro
                table_style.add('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.HexColor('#7f1d1d'))  # Texto rojo oscuro
        
        # Ajustar alineación para las columnas de fecha
        table_style.add('ALIGN', (4, 0), (5, -1), 'CENTER')  # Centrar fechas
        table_style.add('FONTSIZE', (0, 0), (-1, -1), 7)  # Reducir tamaño de fuente para que quepa todo
        
        members_table.setStyle(table_style)
        story.append(members_table)
        
    else:
        story.append(Paragraph("No se encontraron miembros en este semillero.", std_styles['normal']))
    
    story.append(PageBreak())
    
    # Actividades detalladas
    story.append(Paragraph(f"ACTIVIDADES DETALLADAS - {format_semester_label_detailed(semester)}", std_styles['heading']))
    
    if activities:
        # Agrupar por tipo
        activity_types = list(set(a['type'] for a in activities))
        
        for activity_type in activity_types:
            type_activities = [a for a in activities if a['type'] == activity_type]
            
            story.append(Paragraph(f"{activity_type.upper()} ({len(type_activities)})", std_styles['subheading']))
            
            for i, activity in enumerate(type_activities, 1):
                # Información básica - SIN CATEGORÍA
                basic_info = [
                    ['Título:', activity['title'][:50] + '...' if len(activity['title']) > 50 else activity['title']],
                    ['Fecha:', activity['date'].strftime('%d/%m/%Y')],
                    ['Duración:', f"{activity['duration']} horas"],
                    ['Horario:', f"{activity['start_time']} - {activity['end_time']}" if activity['start_time'] and activity['end_time'] else 'No especificado'],
                    ['Horas libres:', 'Aprobadas' if activity['approved_free_hours'] else 'Pendientes']
                ]
                
                # Datos adicionales según disponibilidad
                if activity['reference_number']:
                    basic_info.append(['Número de referencia:', activity['reference_number']])
                
                if activity['publication_date']:
                    basic_info.append(['Fecha publicación:', activity['publication_date'].strftime('%d/%m/%Y')])
                
                if activity['organization_name']:
                    basic_info.append(['Organización:', activity['organization_name']])
                
                # Información específica según el tipo - SIN INVESTIGADORES EN PROYECTOS
                if 'project_data' in activity:
                    project = activity['project_data']
                    basic_info.append(['DATOS DEL PROYECTO', ''])
                    basic_info.extend([
                        ['Nombre proyecto:', project['name'][:40] + '...' if len(project['name']) > 40 else project['name']],
                        ['Ref. proyecto:', project['reference_number']],
                        ['Fecha inicio:', project['start_date'].strftime('%d/%m/%Y') if project['start_date'] else 'N/A'],
                        ['Fecha fin:', project['end_date'].strftime('%d/%m/%Y') if project['end_date'] else 'N/A']
                    ])
                
                elif 'product_data' in activity:
                    product = activity['product_data']
                    basic_info.append(['DATOS DEL PRODUCTO', ''])
                    basic_info.extend([
                        ['Categoría producto:', product['category']],
                        ['Tipo producto:', product['type']],
                        ['Descripción producto:', product['description'][:60] + '...' if len(product['description']) > 60 else product['description']]
                    ])
                
                elif 'recognition_data' in activity:
                    recognition = activity['recognition_data']
                    basic_info.append(['DATOS DEL RECONOCIMIENTO', ''])
                    basic_info.extend([
                        ['Proyecto reconocido:', recognition['project_name'][:40] + '...' if len(recognition['project_name']) > 40 else recognition['project_name']],
                        ['Organización otorgante:', recognition['organization_name'][:40] + '...' if len(recognition['organization_name']) > 40 else recognition['organization_name']]
                    ])
                
                # Autores (MANTENER SOLO ESTAS SECCIONES)
                if activity['authors']['main_authors']:
                    authors_text = ', '.join(activity['authors']['main_authors'])
                    if len(authors_text) > 50:
                        authors_text = authors_text[:50] + '...'
                    basic_info.append(['Autores principales:', authors_text])
                
                if activity['authors']['co_authors']:
                    co_authors_text = ', '.join(activity['authors']['co_authors'])
                    if len(co_authors_text) > 50:
                        co_authors_text = co_authors_text[:50] + '...'
                    basic_info.append(['Co-autores:', co_authors_text])
                
                # Crear tabla de información con estilo especial para separadores
                info_table = Table(basic_info, colWidths=[1.8*inch, 4.7*inch])
                
                # Estilo base
                table_style = get_info_table_style()
                
                # Agregar estilo especial para separadores
                for row_idx, row in enumerate(basic_info):
                    if row[0].startswith('DATOS DEL'):
                        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#e879f9'))  # Rosa medio
                        table_style.add('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white)
                        table_style.add('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
                        table_style.add('ALIGN', (0, row_idx), (-1, row_idx), 'CENTER')
                
                info_table.setStyle(table_style)
                story.append(info_table)
                
                # Descripción completa
                if activity['description']:
                    story.append(Spacer(1, 8))
                    desc_paragraph = Paragraph(f"<b><font color='#be185d'>Descripción:</font></b> {activity['description']}", std_styles['small'])
                    story.append(desc_paragraph)
                
                story.append(Spacer(1, 15))
                
    else:
        no_activities = Paragraph(f"No se encontraron actividades para {format_semester_label_detailed(semester)}.", std_styles['normal'])
        story.append(no_activities)
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def create_projects_sheet(writer, projects, sheet_name):
    """Crea hoja específica para proyectos con formato profesional"""
    projects_data = []
    
    for project in projects:
        project_info = {
            'Título del Informe': project['title'],
            'Fecha de Realización': project['date'].strftime('%d/%m/%Y'),
            'Duración (Horas Académicas)': f"{project['duration']} hrs",
            'Horario de Realización': f"{project['start_time']} - {project['end_time']}" if project['start_time'] and project['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if project['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': project['description'] if project['description'] else 'Sin descripción registrada'
        }
        
        # Datos específicos del proyecto
        if 'project_data' in project:
            project_data = project['project_data']
            project_info.update({
                'Nombre del Proyecto': project_data['name'],
                'Número de Referencia': project_data['reference_number'] if project_data['reference_number'] else 'No asignado',
                'Fecha de Inicio del Proyecto': project_data['start_date'].strftime('%d/%m/%Y') if project_data['start_date'] else 'No registrada',
                'Fecha de Finalización del Proyecto': project_data['end_date'].strftime('%d/%m/%Y') if project_data['end_date'] else 'No registrada'
            })
        
        # Autores con formato profesional
        if project['authors']['main_authors']:
            project_info['Autores Principales'] = '; '.join(project['authors']['main_authors'])
        if project['authors']['co_authors']:
            project_info['Coautores'] = '; '.join(project['authors']['co_authors'])
        
        projects_data.append(project_info)
    
    df_projects = pd.DataFrame(projects_data)
    df_projects.to_excel(writer, sheet_name=sheet_name, index=False)

def create_products_sheet(writer, products, sheet_name):
    """Crea hoja específica para productos con formato profesional"""
    products_data = []
    
    for product in products:
        product_info = {
            'Título del Informe': product['title'],
            'Fecha de Realización': product['date'].strftime('%d/%m/%Y'),
            'Duración (Horas Académicas)': f"{product['duration']} hrs",
            'Horario de Realización': f"{product['start_time']} - {product['end_time']}" if product['start_time'] and product['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if product['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': product['description'] if product['description'] else 'Sin descripción registrada'
        }
        
        # Datos específicos del producto
        if 'product_data' in product:
            product_data = product['product_data']
            product_info.update({
                'Categoría del Producto': product_data['category'],
                'Tipo de Producto': product_data['type'],
                'Descripción del Producto': product_data['description']
            })
        
        # Información adicional
        if product['reference_number']:
            product_info['Número de Referencia'] = product['reference_number']
        if product['publication_date']:
            product_info['Fecha de Publicación'] = product['publication_date'].strftime('%d/%m/%Y')
        if product['organization_name']:
            product_info['Organización/Institución'] = product['organization_name']
        
        # Autores con formato profesional
        if product['authors']['main_authors']:
            product_info['Autores Principales'] = '; '.join(product['authors']['main_authors'])
        if product['authors']['co_authors']:
            product_info['Coautores'] = '; '.join(product['authors']['co_authors'])
        
        products_data.append(product_info)
    
    df_products = pd.DataFrame(products_data)
    df_products.to_excel(writer, sheet_name=sheet_name, index=False)

def create_recognitions_sheet(writer, recognitions, sheet_name):
    """Crea hoja específica para reconocimientos con formato profesional"""
    recognitions_data = []
    
    for recognition in recognitions:
        recognition_info = {
            'Título del Informe': recognition['title'],
            'Fecha de Realización': recognition['date'].strftime('%d/%m/%Y'),
            'Duración (Horas Académicas)': f"{recognition['duration']} hrs",
            'Horario de Realización': f"{recognition['start_time']} - {recognition['end_time']}" if recognition['start_time'] and recognition['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if recognition['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': recognition['description'] if recognition['description'] else 'Sin descripción registrada'
        }
        
        # Datos específicos del reconocimiento
        if 'recognition_data' in recognition:
            recognition_data = recognition['recognition_data']
            recognition_info.update({
                'Proyecto Reconocido': recognition_data['project_name'],
                'Organización Otorgante': recognition_data['organization_name']
            })
        
        # Información adicional
        if recognition['reference_number']:
            recognition_info['Número de Referencia'] = recognition['reference_number']
        if recognition['publication_date']:
            recognition_info['Fecha de Publicación'] = recognition['publication_date'].strftime('%d/%m/%Y')
        if recognition['organization_name']:
            recognition_info['Organización/Institución'] = recognition['organization_name']
        
        # Autores con formato profesional
        if recognition['authors']['main_authors']:
            recognition_info['Autores Principales'] = '; '.join(recognition['authors']['main_authors'])
        if recognition['authors']['co_authors']:
            recognition_info['Coautores'] = '; '.join(recognition['authors']['co_authors'])
        
        recognitions_data.append(recognition_info)
    
    df_recognitions = pd.DataFrame(recognitions_data)
    df_recognitions.to_excel(writer, sheet_name=sheet_name, index=False)

def create_generic_activities_sheet(writer, activities, sheet_name):
    """Crea hoja para actividades genéricas con formato profesional"""
    activities_data = []
    
    for activity in activities:
        activity_info = {
            'Título del Informe': activity['title'],
            'Fecha de Realización': activity['date'].strftime('%d/%m/%Y'),
            'Duración (Horas Académicas)': f"{activity['duration']} hrs",
            'Horario de Realización': f"{activity['start_time']} - {activity['end_time']}" if activity['start_time'] and activity['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if activity['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': activity['description'] if activity['description'] else 'Sin descripción registrada'
        }
        
        # Información adicional si está disponible
        if activity['reference_number']:
            activity_info['Número de Referencia'] = activity['reference_number']
        if activity['publication_date']:
            activity_info['Fecha de Publicación'] = activity['publication_date'].strftime('%d/%m/%Y')
        if activity['organization_name']:
            activity_info['Organización/Institución'] = activity['organization_name']
        
        # Autores con formato profesional
        if activity['authors']['main_authors']:
            activity_info['Autores Principales'] = '; '.join(activity['authors']['main_authors'])
        if activity['authors']['co_authors']:
            activity_info['Coautores'] = '; '.join(activity['authors']['co_authors'])
        
        activities_data.append(activity_info)
    
    df_activities = pd.DataFrame(activities_data)
    df_activities.to_excel(writer, sheet_name=sheet_name, index=False)

def apply_info_sheet_styles(sheet, header_font, subheader_font, body_font, 
                           header_fill, body_fill, alternate_fill, 
                           center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a la hoja de información general"""
    
    # Colores especiales para secciones
    section_blue = '4A90E2'
    section_fill = PatternFill(start_color=section_blue, end_color=section_blue, fill_type='solid')
    section_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            # Headers y títulos de sección
            if cell.value and isinstance(cell.value, str):
                if cell.value.startswith('INFORMACIÓN DEL SEMILLERO') or cell.value.startswith('ESTADÍSTICAS GENERALES'):
                    cell.font = section_font
                    cell.fill = section_fill
                    cell.alignment = center_alignment
                    # Merge cells para títulos de sección
                    if col_num == 1:
                        try:
                            sheet.merge_cells(f'A{row_num}:B{row_num}')
                        except:
                            pass
                elif col_num == 1 and cell.value not in ['', ' ']:  # Primera columna con contenido
                    cell.font = subheader_font
                    cell.fill = body_fill
                    cell.alignment = left_alignment
                else:  # Segunda columna (valores)
                    cell.font = body_font
                    cell.fill = alternate_fill
                    cell.alignment = left_alignment
            else:
                cell.font = body_font
                cell.fill = alternate_fill

def apply_members_sheet_styles(sheet, header_font, body_font, 
                              header_fill, body_fill, alternate_fill,
                              center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a la hoja de miembros"""
    
    # Colores para estados
    active_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Verde claro
    inactive_fill = PatternFill(start_color='FFE8E8', end_color='FFE8E8', fill_type='solid')  # Rojo claro
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            if row_num == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = body_font
                cell.alignment = left_alignment if col_num in [1, 2, 8] else center_alignment  # Nombre, email y observación a la izquierda
                
                # Colorear filas según el estado del miembro
                estado_cell = sheet.cell(row=row_num, column=5)  # Columna "Estado Actual"
                if estado_cell.value == 'Activo':
                    cell.fill = active_fill
                elif estado_cell.value and estado_cell.value != 'Activo':
                    cell.fill = inactive_fill
                else:
                    cell.fill = alternate_fill if row_num % 2 == 0 else body_fill

def apply_activities_sheet_styles(sheet, header_font, body_font,
                                 header_fill, body_fill, alternate_fill,
                                 center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a las hojas de actividades"""
    
    # Color para actividades aprobadas/pendientes
    approved_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Verde claro
    pending_fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')   # Naranja claro
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            if row_num == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = body_font
                
                # Alineación específica por tipo de columna (actualizada para nuevas columnas)
                if col_num in [1, 6]:  # Título del informe y descripción del informe a la izquierda
                    cell.alignment = left_alignment
                elif col_num in [2, 4]:  # Fecha y horario centrados
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
                
                # Colorear filas según estado de aprobación
                # Buscar columna "Estado de Horas Libres"
                horas_libres_col = None
                for col_idx, header_cell in enumerate(sheet[1], 1):
                    if header_cell.value == 'Estado de Horas Libres':
                        horas_libres_col = col_idx
                        break
                
                if horas_libres_col:
                    status_cell = sheet.cell(row=row_num, column=horas_libres_col)
                    if status_cell.value == 'Aprobadas':
                        cell.fill = approved_fill
                    elif status_cell.value == 'Pendientes':
                        cell.fill = pending_fill
                    else:
                        cell.fill = alternate_fill if row_num % 2 == 0 else body_fill
                else:
                    cell.fill = alternate_fill if row_num % 2 == 0 else body_fill