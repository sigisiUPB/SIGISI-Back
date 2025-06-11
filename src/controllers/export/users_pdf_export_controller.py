from flask import jsonify, make_response
from io import BytesIO
from datetime import datetime
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models.users import User
from models.users_research_hotbed import UsersResearchHotbed
from models.research_hotbed import ResearchHotbed
from models.activities_researchHotbed import ActivitiesResearchHotbed
from models.activity_authors import ActivityAuthors
from db.connection import db
from utils.semester_utils import format_semester_label_detailed, is_valid_semester

logger = logging.getLogger(__name__)

def export_user_excel(user_id, semester):
    """
    Exporta los datos de un usuario específico a Excel
    """
    try:
        # Validar semestre
        if not is_valid_semester(semester):
            return jsonify({"error": "Semestre inválido"}), 400
        
        # Buscar el usuario - ARREGLAR deprecated .get()
        user = db.session.get(User, user_id)
        
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404
            
        # Obtener datos del usuario
        research_hotbeds = get_user_research_hotbeds(user_id)
        activities = get_user_activities_by_semester(user_id, semester)
        
        # Generar Excel
        excel_buffer = generate_user_excel_report(user, research_hotbeds, activities, semester)
        
        # Crear respuesta
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{user.idSigaa_user}_{user.name_user.replace(" ", "_")}_{semester}.xlsx"'
        
        logger.info(f"Excel generado exitosamente para usuario {user_id}, semestre {semester}")
        return response
        
    except Exception as e:
        logger.error(f"Error exportando usuario a Excel: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

def export_multiple_users_excel(user_ids, semester):
    """
    Genera archivos Excel individuales para múltiples usuarios
    """
    try:
        # Validaciones
        if not semester or not is_valid_semester(semester):
            return jsonify({"error": "Semestre inválido"}), 400
            
        if not user_ids or len(user_ids) == 0:
            return jsonify({"error": "No se especificaron usuarios"}), 400
            
        # Obtener usuarios
        users = User.query.filter(User.iduser.in_(user_ids)).all()
        if not users:
            return jsonify({"error": "No se encontraron usuarios"}), 404
            
        # Generar Excel consolidado
        excel_buffer = generate_consolidated_users_excel(users, semester)
        
        # Crear respuesta
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="usuarios_consolidado_{semester}.xlsx"'
        
        logger.info(f"Excel consolidado generado exitosamente para {len(users)} usuarios, semestre {semester}")
        return response
        
    except Exception as e:
        logger.error(f"Error generando Excel consolidado: {str(e)}")
        return jsonify({"error": f"Error generando Excel: {str(e)}"}), 500

def generate_user_excel_report(user, research_hotbeds, activities, semester):
    """Genera el reporte Excel individual del usuario con múltiples hojas"""
    buffer = BytesIO()
    
    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # 1. Hoja de información personal
            create_user_info_sheet(writer, user, research_hotbeds, activities, semester)
            
            # 2. Hoja de semilleros asociados
            create_user_hotbeds_sheet(writer, research_hotbeds)
            
            # 3. Hojas de actividades por tipo
            create_user_activities_sheets(writer, activities)
        
        # Recargar el workbook para aplicar estilos
        buffer.seek(0)
        workbook = load_workbook(buffer)
        
        # Verificar que tengamos al menos una hoja
        if len(workbook.sheetnames) == 0:
            ws = workbook.create_sheet("Información Personal")
            ws['A1'] = "No hay datos disponibles"
        
        # Aplicar estilos
        apply_user_excel_styles(workbook)
        
        # Asegurar que al menos una hoja esté visible
        for sheet in workbook.worksheets:
            if sheet.sheet_state == 'hidden':
                sheet.sheet_state = 'visible'
        
        if all(sheet.sheet_state == 'hidden' for sheet in workbook.worksheets):
            workbook.worksheets[0].sheet_state = 'visible'
        
        # Guardar cambios
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        logger.error(f"Error en generate_user_excel_report: {str(e)}")
        # Crear un Excel básico en caso de error
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            error_data = pd.DataFrame({
                'Error': [f'No se pudieron generar los datos: {str(e)}'],
                'Usuario': [user.name_user if user else 'Desconocido'],
                'Semestre': [semester]
            })
            error_data.to_excel(writer, sheet_name='Error', index=False)
        
        buffer.seek(0)
        return buffer

def generate_consolidated_users_excel(users, semester):
    """Genera un Excel consolidado con múltiples usuarios"""
    buffer = BytesIO()
    
    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Crear hoja consolidada de usuarios
            create_consolidated_users_sheet(writer, users, semester)
        
        # Recargar el workbook para aplicar estilos
        buffer.seek(0)
        workbook = load_workbook(buffer)
        
        # Verificar que tengamos al menos una hoja
        if len(workbook.sheetnames) == 0:
            ws = workbook.create_sheet("Usuarios Consolidado")
            ws['A1'] = "No hay datos disponibles"
        
        # Aplicar estilos
        apply_user_excel_styles(workbook)
        
        # Asegurar que al menos una hoja esté visible
        for sheet in workbook.worksheets:
            if sheet.sheet_state == 'hidden':
                sheet.sheet_state = 'visible'
        
        if all(sheet.sheet_state == 'hidden' for sheet in workbook.worksheets):
            workbook.worksheets[0].sheet_state = 'visible'
        
        # Guardar cambios
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        logger.error(f"Error en generate_consolidated_users_excel: {str(e)}")
        # Crear un Excel básico en caso de error
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            error_data = pd.DataFrame({
                'Error': [f'No se pudieron generar los datos: {str(e)}'],
                'Total Usuarios': [len(users) if users else 0],
                'Semestre': [semester]
            })
            error_data.to_excel(writer, sheet_name='Error', index=False)
        
        buffer.seek(0)
        return buffer

def create_user_info_sheet(writer, user, research_hotbeds, activities, semester):
    """Crea la hoja de información personal del usuario"""
    
    # Estadísticas de actividades
    total_hours = sum(a['duration'] for a in activities)
    approved_activities = len([a for a in activities if a['approved_free_hours']])
    
    # Información personal y estadísticas
    info_data = [
        ['INFORMACIÓN PERSONAL', ''],
        ['Nombre Completo', user.name_user],
        ['ID SIGAA', user.idSigaa_user or 'No registrado'],
        ['Correo Electrónico', user.email_user],
        ['Programa Académico', user.academicProgram_user or 'No especificado'],
        ['Tipo de Usuario', user.type_user],
        ['Estado del Usuario', user.status_user],
        ['Período del Reporte', format_semester_label_detailed(semester)],
        ['Fecha de Generación', datetime.now().strftime('%d de %B de %Y - %H:%M hrs')],
        ['', ''],
        ['ESTADÍSTICAS DE ACTIVIDADES', ''],
        ['Total de Semilleros Asociados', f'{len(research_hotbeds)} semilleros'],
        ['Semilleros Activos', f'{len([h for h in research_hotbeds if h["status"] == "Activo"])} semilleros'],
        ['Total de Actividades', f'{len(activities)} actividades'],
        ['Horas Académicas Totales', f'{total_hours} horas'],
        ['Actividades Aprobadas', f'{approved_activities} actividades'],
        ['Actividades Pendientes', f'{len(activities) - approved_activities} actividades'],
        ['Proyectos de Investigación', f'{len([a for a in activities if "proyecto" in a["type"].lower()])} proyectos'],
        ['Productos Académicos', f'{len([a for a in activities if "producto" in a["type"].lower()])} productos'],
        ['Reconocimientos', f'{len([a for a in activities if "reconocimiento" in a["type"].lower()])} reconocimientos']
    ]
    
    df_info = pd.DataFrame(info_data, columns=['Concepto', 'Detalle'])
    df_info.to_excel(writer, sheet_name='Información Personal', index=False)

def create_user_hotbeds_sheet(writer, research_hotbeds):
    """Crea la hoja de semilleros asociados al usuario"""
    if not research_hotbeds:
        df_empty = pd.DataFrame({'Información': ['No está asociado a ningún semillero de investigación.']})
        df_empty.to_excel(writer, sheet_name='Semilleros', index=False)
        return
    
    # Separar semilleros activos e inactivos
    active_hotbeds = [h for h in research_hotbeds if h['status'] == 'Activo']
    inactive_hotbeds = [h for h in research_hotbeds if h['status'] != 'Activo']
    sorted_hotbeds = active_hotbeds + inactive_hotbeds
    
    hotbeds_data = []
    for hotbed in sorted_hotbeds:
        hotbeds_data.append({
            'Nombre del Semillero': hotbed['name'],
            'Acrónimo/Siglas': hotbed['acronym'],
            'Facultad': hotbed['faculty'],
            'Sede Universitaria': hotbed['universityBranch'],
            'Rol en el Semillero': hotbed['type'],
            'Estado de Participación': hotbed['status'],
            'Fecha de Ingreso': hotbed['dateEnter'].strftime('%d/%m/%Y') if hotbed['dateEnter'] else 'No registrada'
        })
    
    df_hotbeds = pd.DataFrame(hotbeds_data)
    df_hotbeds.to_excel(writer, sheet_name='Semilleros', index=False)

def create_user_activities_sheets(writer, activities):
    """Crea hojas separadas para cada tipo de actividad del usuario"""
    if not activities:
        df_empty = pd.DataFrame({'Mensaje': ['No hay actividades registradas para este semestre.']})
        df_empty.to_excel(writer, sheet_name='Actividades', index=False)
        return
    
    # Agrupar actividades por tipo
    activities_by_type = {}
    for activity in activities:
        activity_type = activity['type']
        if activity_type not in activities_by_type:
            activities_by_type[activity_type] = []
        activities_by_type[activity_type].append(activity)
    
    # Crear hoja para cada tipo
    for activity_type, type_activities in activities_by_type.items():
        sheet_name = get_user_sheet_name_for_type(activity_type)
        
        if 'proyecto' in activity_type.lower():
            create_user_projects_sheet(writer, type_activities, sheet_name)
        elif 'producto' in activity_type.lower():
            create_user_products_sheet(writer, type_activities, sheet_name)
        elif 'reconocimiento' in activity_type.lower():
            create_user_recognitions_sheet(writer, type_activities, sheet_name)
        else:
            create_user_generic_activities_sheet(writer, type_activities, sheet_name)

def create_consolidated_users_sheet(writer, users, semester):
    """Crea la hoja consolidada de usuarios"""
    
    # Información general del reporte
    info_data = [
        ['REPORTE CONSOLIDADO DE USUARIOS', ''],
        ['Período del Reporte', format_semester_label_detailed(semester)],
        ['Total de Usuarios', f'{len(users)} usuarios'],
        ['Fecha de Generación', datetime.now().strftime('%d de %B de %Y - %H:%M hrs')],
        ['Sistema', 'SIGISI - Sistema de Gestión de Semilleros'],
        ['', '']
    ]
    
    # Datos consolidados de usuarios
    users_data = []
    for i, user in enumerate(users, 1):
        activities = get_user_activities_by_semester(user.iduser, semester)
        research_hotbeds = get_user_research_hotbeds(user.iduser)
        total_hours = sum(a['duration'] for a in activities)
        approved_activities = len([a for a in activities if a['approved_free_hours']])
        
        users_data.append({
            'No.': i,
            'Nombre Completo': user.name_user,
            'ID SIGAA': user.idSigaa_user or 'No registrado',
            'Correo Electrónico': user.email_user,
            'Tipo de Usuario': user.type_user,
            'Estado': user.status_user,
            'Semilleros Asociados': len(research_hotbeds),
            'Total Actividades': len(activities),
            'Horas Totales': total_hours,
            'Actividades Aprobadas': approved_activities,
            'Actividades Pendientes': len(activities) - approved_activities,
            'Proyectos': len([a for a in activities if 'proyecto' in a['type'].lower()]),
            'Productos': len([a for a in activities if 'producto' in a['type'].lower()]),
            'Reconocimientos': len([a for a in activities if 'reconocimiento' in a['type'].lower()])
        })
    
    # Crear DataFrame combinado
    df_info = pd.DataFrame(info_data, columns=['Concepto', 'Detalle'])
    df_users = pd.DataFrame(users_data)
    
    # Combinar ambos DataFrames con espacios
    combined_data = []
    
    # Agregar información general
    for _, row in df_info.iterrows():
        combined_data.append([row['Concepto'], row['Detalle']] + [''] * (len(df_users.columns) - 2))
    
    # Agregar espacio
    combined_data.append([''] * len(df_users.columns))
    combined_data.append(['LISTADO DETALLADO DE USUARIOS'] + [''] * (len(df_users.columns) - 1))
    combined_data.append([''] * len(df_users.columns))
    
    # Agregar headers de usuarios
    combined_data.append(df_users.columns.tolist())
    
    # Agregar datos de usuarios
    for _, row in df_users.iterrows():
        combined_data.append(row.tolist())
    
    # Crear DataFrame final y guardar
    df_final = pd.DataFrame(combined_data)
    df_final.to_excel(writer, sheet_name='Usuarios Consolidado', index=False, header=False)

def get_user_sheet_name_for_type(activity_type):
    """Obtiene el nombre de la hoja según el tipo de actividad para usuarios"""
    if 'proyecto' in activity_type.lower():
        return 'Proyectos'
    elif 'producto' in activity_type.lower():
        return 'Productos'
    elif 'reconocimiento' in activity_type.lower():
        return 'Reconocimientos'
    else:
        # Limpiar el nombre y limitar a 31 caracteres (límite de Excel)
        clean_name = activity_type.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        return clean_name.capitalize()[:31]

def create_user_projects_sheet(writer, projects, sheet_name):
    """Crea hoja específica para proyectos del usuario"""
    projects_data = []
    
    for project in projects:
        project_info = {
            'Título del Informe': project['title'],
            'Fecha de Realización': project['date'].strftime('%d/%m/%Y'),
            'Semillero Asociado': project['research_hotbed_name'],
            'Duración (Horas Académicas)': f"{project['duration']} hrs",
            'Horario de Realización': f"{project['start_time']} - {project['end_time']}" if project['start_time'] and project['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if project['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': project['description'] if project['description'] else 'Sin descripción registrada'
        }
        
        # Datos específicos del proyecto
        if 'project_data' in project:
            project_data = project['project_data']
            project_info.update({
                'Nombre del Proyecto': project_data['name'],
                'Número de Referencia': project_data['reference_number'] if project_data['reference_number'] else 'No asignado',
                'Fecha de Inicio del Proyecto': project_data['start_date'].strftime('%d/%m/%Y') if project_data['start_date'] else 'No registrada',
                'Fecha de Finalización del Proyecto': project_data['end_date'].strftime('%d/%m/%Y') if project_data['end_date'] else 'No registrada'
            })
        
        # Autores
        if project['authors']['main_authors']:
            project_info['Autores Principales'] = '; '.join(project['authors']['main_authors'])
        if project['authors']['co_authors']:
            project_info['Coautores'] = '; '.join(project['authors']['co_authors'])
        
        projects_data.append(project_info)
    
    df_projects = pd.DataFrame(projects_data)
    df_projects.to_excel(writer, sheet_name=sheet_name, index=False)

def create_user_products_sheet(writer, products, sheet_name):
    """Crea hoja específica para productos del usuario"""
    products_data = []
    
    for product in products:
        product_info = {
            'Título del Informe': product['title'],
            'Fecha de Realización': product['date'].strftime('%d/%m/%Y'),
            'Semillero Asociado': product['research_hotbed_name'],
            'Duración (Horas Académicas)': f"{product['duration']} hrs",
            'Horario de Realización': f"{product['start_time']} - {product['end_time']}" if product['start_time'] and product['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if product['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': product['description'] if product['description'] else 'Sin descripción registrada'
        }
        
        # Datos específicos del producto
        if 'product_data' in product:
            product_data = product['product_data']
            product_info.update({
                'Categoría del Producto': product_data['category'],
                'Tipo de Producto': product_data['type'],
                'Descripción del Producto': product_data['description']
            })
        
        # Información adicional
        if product['reference_number']:
            product_info['Número de Referencia'] = product['reference_number']
        if product['publication_date']:
            product_info['Fecha de Publicación'] = product['publication_date'].strftime('%d/%m/%Y')
        if product['organization_name']:
            product_info['Organización/Institución'] = product['organization_name']
        
        # Autores
        if product['authors']['main_authors']:
            product_info['Autores Principales'] = '; '.join(product['authors']['main_authors'])
        if product['authors']['co_authors']:
            product_info['Coautores'] = '; '.join(product['authors']['co_authors'])
        
        products_data.append(product_info)
    
    df_products = pd.DataFrame(products_data)
    df_products.to_excel(writer, sheet_name=sheet_name, index=False)

def create_user_recognitions_sheet(writer, recognitions, sheet_name):
    """Crea hoja específica para reconocimientos del usuario"""
    recognitions_data = []
    
    for recognition in recognitions:
        recognition_info = {
            'Título del Informe': recognition['title'],
            'Fecha de Realización': recognition['date'].strftime('%d/%m/%Y'),
            'Semillero Asociado': recognition['research_hotbed_name'],
            'Duración (Horas Académicas)': f"{recognition['duration']} hrs",
            'Horario de Realización': f"{recognition['start_time']} - {recognition['end_time']}" if recognition['start_time'] and recognition['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if recognition['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': recognition['description'] if recognition['description'] else 'Sin descripción registrada'
        }
        
        # Datos específicos del reconocimiento
        if 'recognition_data' in recognition:
            recognition_data = recognition['recognition_data']
            recognition_info.update({
                'Proyecto Reconocido': recognition_data['project_name'],
                'Organización Otorgante': recognition_data['organization_name']
            })
        
        # Información adicional
        if recognition['reference_number']:
            recognition_info['Número de Referencia'] = recognition['reference_number']
        if recognition['publication_date']:
            recognition_info['Fecha de Publicación'] = recognition['publication_date'].strftime('%d/%m/%Y')
        if recognition['organization_name']:
            recognition_info['Organización/Institución'] = recognition['organization_name']
        
        # Autores
        if recognition['authors']['main_authors']:
            recognition_info['Autores Principales'] = '; '.join(recognition['authors']['main_authors'])
        if recognition['authors']['co_authors']:
            recognition_info['Coautores'] = '; '.join(recognition['authors']['co_authors'])
        
        recognitions_data.append(recognition_info)
    
    df_recognitions = pd.DataFrame(recognitions_data)
    df_recognitions.to_excel(writer, sheet_name=sheet_name, index=False)

def create_user_generic_activities_sheet(writer, activities, sheet_name):
    """Crea hoja para actividades genéricas del usuario"""
    activities_data = []
    
    for activity in activities:
        activity_info = {
            'Título del Informe': activity['title'],
            'Fecha de Realización': activity['date'].strftime('%d/%m/%Y'),
            'Semillero Asociado': activity['research_hotbed_name'],
            'Duración (Horas Académicas)': f"{activity['duration']} hrs",
            'Horario de Realización': f"{activity['start_time']} - {activity['end_time']}" if activity['start_time'] and activity['end_time'] else 'No especificado',
            'Estado de Horas Libres': 'Aprobadas' if activity['approved_free_hours'] else 'Pendientes',
            'Descripción del Informe': activity['description'] if activity['description'] else 'Sin descripción registrada'
        }
        
        # Información adicional si está disponible
        if activity['reference_number']:
            activity_info['Número de Referencia'] = activity['reference_number']
        if activity['publication_date']:
            activity_info['Fecha de Publicación'] = activity['publication_date'].strftime('%d/%m/%Y')
        if activity['organization_name']:
            activity_info['Organización/Institución'] = activity['organization_name']
        
        # Autores
        if activity['authors']['main_authors']:
            activity_info['Autores Principales'] = '; '.join(activity['authors']['main_authors'])
        if activity['authors']['co_authors']:
            activity_info['Coautores'] = '; '.join(activity['authors']['co_authors'])
        
        activities_data.append(activity_info)
    
    df_activities = pd.DataFrame(activities_data)
    df_activities.to_excel(writer, sheet_name=sheet_name, index=False)

def apply_user_excel_styles(workbook):
    """Aplica estilos profesionales para usuarios al workbook"""
    
    try:
        # Colores formales para usuarios (tonos verdes y azules)
        primary_green = '1E7B4F'     # Verde corporativo
        secondary_green = '22C55E'   # Verde medio
        light_green = 'ECFDF5'       # Verde muy claro
        primary_blue = '1F4E79'      # Azul marino
        dark_gray = '2F2F2F'         # Gris oscuro
        medium_gray = '5A5A5A'       # Gris medio
        light_gray = 'F8F9FA'        # Gris muy claro
        white = 'FFFFFF'             # Blanco
        
        # Estilos profesionales
        header_font = Font(bold=True, color=white, size=11, name='Calibri')
        subheader_font = Font(bold=True, color=dark_gray, size=10, name='Calibri')
        body_font = Font(color=dark_gray, size=10, name='Calibri')
        
        # Fills (fondos)
        header_fill = PatternFill(start_color=primary_green, end_color=primary_green, fill_type='solid')
        subheader_fill = PatternFill(start_color=secondary_green, end_color=secondary_green, fill_type='solid')
        body_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type='solid')
        alternate_fill = PatternFill(start_color=white, end_color=white, fill_type='solid')
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Bordes profesionales
        thin_border = Side(border_style='thin', color=medium_gray)
        medium_border = Side(border_style='medium', color=primary_green)
        
        header_border = Border(
            left=thin_border, right=thin_border, 
            top=medium_border, bottom=medium_border
        )
        
        body_border = Border(
            left=thin_border, right=thin_border, 
            top=thin_border, bottom=thin_border
        )
        
        # Aplicar estilos a todas las hojas
        for sheet_name in workbook.sheetnames:
            try:
                sheet = workbook[sheet_name]
                
                # Asegurar que la hoja esté visible
                sheet.sheet_state = 'visible'
                
                # Configurar ancho de columnas
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = max(len(str(line)) for line in str(cell.value).split('\n'))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # Ajustar ancho
                    if max_length == 0:
                        adjusted_width = 12
                    elif max_length < 8:
                        adjusted_width = 12
                    elif max_length < 20:
                        adjusted_width = max_length + 3
                    elif max_length < 40:
                        adjusted_width = max_length + 2
                    else:
                        adjusted_width = 45
                        
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar estilos específicos según el tipo de hoja
                if sheet_name == 'Información Personal':
                    apply_user_info_sheet_styles(sheet, header_font, subheader_font, body_font, 
                                                header_fill, body_fill, alternate_fill, 
                                                center_alignment, left_alignment, header_border, body_border)
                
                elif sheet_name == 'Semilleros':
                    apply_user_hotbeds_sheet_styles(sheet, header_font, body_font, 
                                                   header_fill, body_fill, alternate_fill,
                                                   center_alignment, left_alignment, header_border, body_border)
                
                elif sheet_name == 'Usuarios Consolidado':
                    apply_consolidated_sheet_styles(sheet, header_font, subheader_font, body_font, 
                                                  header_fill, body_fill, alternate_fill,
                                                  center_alignment, left_alignment, header_border, body_border)
                
                else:  # Hojas de actividades
                    apply_user_activities_sheet_styles(sheet, header_font, body_font,
                                                      header_fill, body_fill, alternate_fill,
                                                      center_alignment, left_alignment, header_border, body_border)
            except Exception as e:
                logger.error(f"Error aplicando estilos a la hoja {sheet_name}: {str(e)}")
                continue
                
    except Exception as e:
        logger.error(f"Error general aplicando estilos: {str(e)}")

def apply_user_info_sheet_styles(sheet, header_font, subheader_font, body_font, 
                                header_fill, body_fill, alternate_fill, 
                                center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a la hoja de información personal"""
    
    # Colores especiales para secciones de usuario
    section_green = '16A34A'
    section_fill = PatternFill(start_color=section_green, end_color=section_green, fill_type='solid')
    section_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            # Headers y títulos de sección
            if cell.value and isinstance(cell.value, str):
                if cell.value.startswith('INFORMACIÓN PERSONAL') or cell.value.startswith('ESTADÍSTICAS DE ACTIVIDADES'):
                    cell.font = section_font
                    cell.fill = section_fill
                    cell.alignment = center_alignment
                    # Merge cells para títulos de sección
                    if col_num == 1:
                        try:
                            sheet.merge_cells(f'A{row_num}:B{row_num}')
                        except:
                            pass
                elif col_num == 1 and cell.value not in ['', ' ']:  # Primera columna con contenido
                    cell.font = subheader_font
                    cell.fill = body_fill
                    cell.alignment = left_alignment
                else:  # Segunda columna (valores)
                    cell.font = body_font
                    cell.fill = alternate_fill
                    cell.alignment = left_alignment
            else:
                cell.font = body_font
                cell.fill = alternate_fill

def apply_user_hotbeds_sheet_styles(sheet, header_font, body_font, 
                                   header_fill, body_fill, alternate_fill,
                                   center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a la hoja de semilleros del usuario"""
    
    # Colores para estados de participación
    active_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')  # Verde claro
    inactive_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')  # Rojo claro
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            if row_num == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = body_font
                cell.alignment = left_alignment if col_num in [1, 3, 4] else center_alignment  # Nombre, facultad, sede a la izquierda
                
                # Colorear filas según el estado de participación
                estado_cell = sheet.cell(row=row_num, column=6)  # Columna "Estado de Participación"
                if estado_cell.value == 'Activo':
                    cell.fill = active_fill
                elif estado_cell.value and estado_cell.value != 'Activo':
                    cell.fill = inactive_fill
                else:
                    cell.fill = alternate_fill if row_num % 2 == 0 else body_fill

def apply_user_activities_sheet_styles(sheet, header_font, body_font,
                                      header_fill, body_fill, alternate_fill,
                                      center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a las hojas de actividades del usuario"""
    
    # Color para actividades aprobadas/pendientes
    approved_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')  # Verde claro
    pending_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')   # Amarillo claro
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            if row_num == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = body_font
                
                # Alineación específica por tipo de columna
                if col_num in [1, 3, 7]:  # Título, semillero, descripción a la izquierda
                    cell.alignment = left_alignment
                elif col_num in [2, 4, 5]:  # Fecha, duración, horario centrados
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
                
                # Colorear filas según estado de horas libres
                # Buscar columna "Estado de Horas Libres"
                horas_libres_col = None
                
                for col_idx, header_cell in enumerate(sheet[1], 1):
                    if header_cell.value == 'Estado de Horas Libres':
                        horas_libres_col = col_idx
                        break
                
                if horas_libres_col:
                    status_cell = sheet.cell(row=row_num, column=horas_libres_col)
                    if status_cell.value == 'Aprobadas':
                        cell.fill = approved_fill
                    elif status_cell.value == 'Pendientes':
                        cell.fill = pending_fill
                    else:
                        cell.fill = alternate_fill if row_num % 2 == 0 else body_fill
                else:
                    cell.fill = alternate_fill if row_num % 2 == 0 else body_fill

def apply_consolidated_sheet_styles(sheet, header_font, subheader_font, body_font, 
                                   header_fill, body_fill, alternate_fill,
                                   center_alignment, left_alignment, header_border, body_border):
    """Aplica estilos específicos a la hoja consolidada"""
    
    # Color especial para títulos
    title_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        for col_num, cell in enumerate(row, 1):
            cell.border = body_border
            
            if cell.value and isinstance(cell.value, str):
                if 'REPORTE CONSOLIDADO' in cell.value or 'LISTADO DETALLADO' in cell.value:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = center_alignment
                elif cell.value in ['No.', 'Nombre Completo', 'ID SIGAA', 'Correo Electrónico']:  # Headers de tabla
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                elif col_num == 1 and cell.value not in ['', ' ', 'No.']:  # Primera columna info general
                    cell.font = subheader_font
                    cell.fill = body_fill
                    cell.alignment = left_alignment
                else:
                    cell.font = body_font
                    cell.fill = alternate_fill if row_num % 2 == 0 else body_fill
                    cell.alignment = center_alignment if str(cell.value).isdigit() else left_alignment
            else:
                cell.font = body_font
                cell.fill = alternate_fill

# Funciones auxiliares que ya existían
def get_user_research_hotbeds(user_id):
    """Obtiene los semilleros asociados al usuario"""
    try:
        hotbeds_query = db.session.query(ResearchHotbed, UsersResearchHotbed).join(
            UsersResearchHotbed, ResearchHotbed.idresearchHotbed == UsersResearchHotbed.researchHotbed_idresearchHotbed
        ).filter(
            UsersResearchHotbed.user_iduser == user_id
        ).all()
        
        return [{
            'name': hotbed.name_researchHotbed,
            'acronym': hotbed.acronym_researchHotbed,
            'faculty': hotbed.faculty_researchHotbed,
            'universityBranch': hotbed.universityBranch_researchHotbed,
            'type': user_research.TypeUser_usersResearchHotbed,
            'status': user_research.status_usersResearchHotbed,
            'dateEnter': user_research.dateEnter_usersResearchHotbed
        } for hotbed, user_research in hotbeds_query]
        
    except Exception as e:
        logger.error(f"Error obteniendo semilleros del usuario: {str(e)}")
        return []

def get_user_activities_by_semester(user_id, semester):
    """Obtiene actividades del usuario filtradas por semestre - INCLUYE ACTIVIDADES COMO CO-AUTOR"""
    try:
        # Obtener IDs de semilleros del usuario
        user_research_ids = db.session.query(UsersResearchHotbed.idusersResearchHotbed).filter(
            UsersResearchHotbed.user_iduser == user_id
        ).subquery()
        
        # OPCIÓN 1: Actividades donde el usuario está directamente asociado al semillero
        direct_activities = db.session.query(ActivitiesResearchHotbed).filter(
            ActivitiesResearchHotbed.usersResearchHotbed_idusersResearchHotbed.in_(user_research_ids),
            ActivitiesResearchHotbed.semester == semester
        )
        
        # OPCIÓN 2: Actividades donde el usuario es autor/co-autor (a través de ActivityAuthors)
        authored_activities = db.session.query(ActivitiesResearchHotbed).join(
            ActivityAuthors, ActivitiesResearchHotbed.idactivitiesResearchHotbed == ActivityAuthors.activity_id
        ).filter(
            ActivityAuthors.user_research_hotbed_id.in_(user_research_ids),
            ActivitiesResearchHotbed.semester == semester
        )
        
        # Combinar ambas consultas y eliminar duplicados
        all_activities_query = direct_activities.union(authored_activities).all()
        
        # Procesar actividades
        filtered_activities = []
        processed_ids = set()  # Para evitar duplicados
        
        for activity in all_activities_query:
            # Evitar duplicados
            if activity.idactivitiesResearchHotbed in processed_ids:
                continue
            processed_ids.add(activity.idactivitiesResearchHotbed)
            
            # Obtener autores
            authors_data = get_activity_authors(activity.idactivitiesResearchHotbed)
            
            # Obtener nombre del semillero
            research_hotbed_name = get_research_hotbed_name(activity.usersResearchHotbed_idusersResearchHotbed)
            
            activity_data = {
                'id': activity.idactivitiesResearchHotbed,
                'title': activity.title_activitiesResearchHotbed,
                'responsible': activity.responsible_activitiesResearchHotbed,
                'date': activity.date_activitiesResearchHotbed,
                'description': activity.description_activitiesResearchHotbed,
                'type': activity.type_activitiesResearchHotbed,
                'duration': activity.duration_activitiesResearchHotbed or 0,
                'start_time': activity.startTime_activitiesResearchHotbed,
                'end_time': activity.endTime_activitiesResearchHotbed,
                'approved_free_hours': bool(activity.approvedFreeHours_activitiesResearchHotbed),
                'reference_number': getattr(activity, 'reference_number', None),
                'publication_date': getattr(activity, 'publication_date', None),
                'organization_name': getattr(activity, 'organization_name', None),
                'authors': authors_data,
                'research_hotbed_name': research_hotbed_name
            }
            
            # Obtener datos específicos según el tipo
            if (activity.type_activitiesResearchHotbed and 
                activity.type_activitiesResearchHotbed.lower() == 'proyecto' and 
                activity.projectsResearchHotbed_idprojectsResearchHotbed):
                
                from models.projects_researchHotbed import ProjectsResearchHotbed
                project = db.session.query(ProjectsResearchHotbed).filter_by(
                    idprojectsResearchHotbed=activity.projectsResearchHotbed_idprojectsResearchHotbed
                ).first()
                if project:
                    activity_data['project_data'] = {
                        'name': project.name_projectsResearchHotbed or 'Sin especificar',
                        'reference_number': project.referenceNumber_projectsResearchHotbed,
                        'start_date': project.startDate_projectsResearchHotbed,
                        'end_date': project.endDate_projectsResearchHotbed
                    }
            
            elif (activity.type_activitiesResearchHotbed and 
                  activity.type_activitiesResearchHotbed.lower() == 'producto' and 
                  activity.productsResearchHotbed_idproductsResearchHotbed):
                
                from models.products_researchHotbed import ProductsResearchHotbed
                product = db.session.query(ProductsResearchHotbed).filter_by(
                    idproductsResearchHotbed=activity.productsResearchHotbed_idproductsResearchHotbed
                ).first()
                if product:
                    activity_data['product_data'] = {
                        'category': product.category_productsResearchHotbed,
                        'type': product.type_productsResearchHotbed,
                        'description': product.description_productsResearchHotbed
                    }
            
            elif (activity.type_activitiesResearchHotbed and 
                  activity.type_activitiesResearchHotbed.lower() == 'reconocimiento' and 
                  activity.recognitionsResearchHotbed_idrecognitionsResearchHotbed):
                
                from models.recognitions_researchHotbed import RecognitionsResearchHotbed
                recognition = db.session.query(RecognitionsResearchHotbed).filter_by(
                    idrecognitionsResearchHotbed=activity.recognitionsResearchHotbed_idrecognitionsResearchHotbed
                ).first()
                if recognition:
                    activity_data['recognition_data'] = {
                        'project_name': recognition.projectName_recognitionsResearchHotbed,
                        'organization_name': recognition.organizationName_recognitionsResearchHotbed
                    }
            
            filtered_activities.append(activity_data)
                
        return filtered_activities
        
    except Exception as e:
        logger.error(f"Error obteniendo actividades del usuario: {str(e)}")
        return []

def get_activity_authors(activity_id):
    """Obtiene los autores de una actividad"""
    try:
        authors_query = db.session.query(User, ActivityAuthors).join(
            UsersResearchHotbed, ActivityAuthors.user_research_hotbed_id == UsersResearchHotbed.idusersResearchHotbed
        ).join(
            User, UsersResearchHotbed.user_iduser == User.iduser
        ).filter(
            ActivityAuthors.activity_id == activity_id
        ).all()
        
        main_authors = []
        co_authors = []
        
        for user, author_relation in authors_query:
            if author_relation.is_main_author:
                main_authors.append(user.name_user)
            else:
                co_authors.append(user.name_user)
                
        return {
            'main_authors': main_authors,
            'co_authors': co_authors
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo autores: {str(e)}")
        return {'main_authors': [], 'co_authors': []}

def get_research_hotbed_name(user_research_hotbed_id):
    """Obtiene el nombre del semillero por ID de user_research_hotbed"""
    try:
        result = db.session.query(ResearchHotbed.name_researchHotbed).join(
            UsersResearchHotbed, ResearchHotbed.idresearchHotbed == UsersResearchHotbed.researchHotbed_idresearchHotbed
        ).filter(
            UsersResearchHotbed.idusersResearchHotbed == user_research_hotbed_id
        ).first()
        
        return result[0] if result else 'Semillero no especificado'
        
    except Exception as e:
        logger.error(f"Error obteniendo nombre del semillero: {str(e)}")
        return 'Semillero no especificado'

def get_user_role_in_activity_simple(activity):
    """Determina el rol del usuario en una actividad de forma simplificada"""
    # Esta función ya no se usa, pero se mantiene por compatibilidad
    return "Participante"

def get_activities_by_semester(research_hotbed_id, semester):
    """Obtiene actividades filtradas por semestre usando el campo 'semester' de la actividad"""
    try:
        activities_query = db.session.query(ActivitiesResearchHotbed).join(
            UsersResearchHotbed, ActivitiesResearchHotbed.usersResearchHotbed_idusersResearchHotbed == UsersResearchHotbed.idusersResearchHotbed
        ).filter(
            UsersResearchHotbed.researchHotbed_idresearchHotbed == research_hotbed_id,
            ActivitiesResearchHotbed.semester == semester
        ).all()
        
        # Procesar actividades similar a get_user_activities_by_semester
        filtered_activities = []
        
        for activity in activities_query:
            # Obtener autores
            authors_data = get_activity_authors(activity.idactivitiesResearchHotbed)
            
            # Obtener nombre del semillero
            research_hotbed_name = get_research_hotbed_name(activity.usersResearchHotbed_idusersResearchHotbed)
            
            activity_data = {
                'id': activity.idactivitiesResearchHotbed,
                'title': activity.title_activitiesResearchHotbed,
                'responsible': activity.responsible_activitiesResearchHotbed,
                'date': activity.date_activitiesResearchHotbed,
                'description': activity.description_activitiesResearchHotbed,
                'type': activity.type_activitiesResearchHotbed,
                'duration': activity.duration_activitiesResearchHotbed or 0,
                'start_time': activity.startTime_activitiesResearchHotbed,
                'end_time': activity.endTime_activitiesResearchHotbed,
                'approved_free_hours': bool(activity.approvedFreeHours_activitiesResearchHotbed),
                'reference_number': getattr(activity, 'reference_number', None),
                'publication_date': getattr(activity, 'publication_date', None),
                'organization_name': getattr(activity, 'organization_name', None),
                'authors': authors_data,
                'research_hotbed_name': research_hotbed_name
            }
            
            filtered_activities.append(activity_data)
                
        return filtered_activities
        
    except Exception as e:
        logger.error(f"Error obteniendo actividades: {str(e)}")
        return []

# Funciones exportadas para mantener compatibilidad
export_user_pdf = export_user_excel  # Alias para compatibilidad
export_multiple_users_pdf = export_multiple_users_excel  # Alias para compatibilidad